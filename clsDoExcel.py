from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
import os

class DoExcel:
    
    def __init__(self,folderPath):
        ''' Sets Folder path,,alphabet dictionary and creates empty dictionary for placeholder '''

        self.fPath=folderPath
        self.mappedCodes={}
        self.__alpha={1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M',\
               14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}

##-----------------------------------------------------------------------------------------------------------------------------------
    def loadFile(self,fileName):
        ''' Opens excel file with active sheet to work on.
            returns:- workbook as book,worksheet as sheet,cell,lastColumn,lastRow'''
        
        self.file=fileName
        self.book=load_workbook(os.path.join(self.fPath,self.file))
        self.sheet=self.book.active
        self.cell=self.sheet.cell
        self.lastColumn=self.sheet.max_column
        self.lastRow=self.sheet.max_row
        return self.file

##-----------------------------------------------------------------------------------------------------------------------------------
    def displayColumns(self,row=1,blncolNum=False):
        ''' Displays column names in with headers and returns a dictionary for the same'''

        self.row=row
        for col in range(1,self.lastColumn+1):
            print("{} : {}".format(col,self.sheet.cell(row=self.row,column=col).value))
            self.mappedCodes[col]=self.sheet.cell(row=self.row,column=col).value
        return self.mappedCodes

##-----------------------------------------------------------------------------------------------------------------------------------
    def writeColumnHeaders(self,row=1,*args):
        '''Accepts row number and strings of headers and writes to the columns in specified row number
            returns list of headers'''

        self.row=row
        self.lstNewHeader=[]
        for header in args:
            self.writeToCell(self.row,self.sheet.max_column+1,header)
            self.lstNewHeader.append(self.sheet.max_column)
        return self.lstNewHeader

##-----------------------------------------------------------------------------------------------------------------------------------
    def mapColsList(self,row=1,lstCols=[]):
        ''' Accepts in column numbers in list and returns consecutive Column in alphabets'''

        self.row=row
        self.lstCols=lstCols
        self.lstMappedCols=[]
        for item in lstCols:
            if not item==",":
                item=int(item)
                self.lstMappedCols.append(self.mapColumn(item,row))
        return self.lstMappedCols

##-----------------------------------------------------------------------------------------------------------------------------------
    def mapColumn(self,num,row=None,flgColNum=False):
        ''' Accepts in number,optional row number and boolean for columnnumber - returns corresponding column name (Private method)'''
        
        self.num=num
        self.row=row
        self.blnColNum=flgColNum
        if self.blnColNum:
            self.colName=str(self.__alpha.get(int(self.num),0))+str(self.row)
        else:
            self.colName=str(self.__alpha.get(int(self.num),0))
        return self.colName

##----------------------------------------------------------------------------------------------------------------------------------
    def mapNumToCol(self,col):
        ''' Accepts Alphabet and maps column number, return consecutive column number '''

        for k,v in self.__alpha.items():
            if v==col:
                return k      

##-----------------------------------------------------------------------------------------------------------------------------------
    def deleteCellData(self, row,col):
        ''' Deletes data in cell mentioned by row and column number'''

        self.row=row
        self.col=col
        self.sheet.cell(row=self.row,column=self.col).value=None

##-----------------------------------------------------------------------------------------------------------------------------------
    def writeToCell(self,row,col,value):
        ''' Writes to cell, arguments row number,column number and data to be written'''

        self.row=row
        self.col=col
        self.value=value
        self.sheet.cell(row=self.row,column=self.col).value=self.value

##-----------------------------------------------------------------------------------------------------------------------------------
    def writeFormula(self,rowNum,colNum,strFormula):
        '''Accepts row number, column number, string for formula and writes it to specified cell'''

        pass
##----------------------------------------------------------------------------------------------------------------------------------
    def highlightCell(self,cell):
        ''' Accepts in cell address and fills the background to yellow color '''

        self.cell=cell
        self.sheet[self.cell].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
##---------------------------------------------------------------------------------------------------------------------------------
    def highlightRow(self,row):
        ''' Accepts in row number and fills the background to yellow color '''

        self.row=row+1
        for i in range(1,self.lastColumn):
            self.cell=str(self.mapColumn(i,self.row,True))
            self.highlightCell(self.cell)
####----------------------------------------------------------------------------------------------------------------------------------
    def fetchValue(self,rowNum,col):
        ''' Accepts in row column and returns value from cell '''
        
        cellVal=self.sheet[col][rowNum].value

        return cellVal
##---------------------------------------------------------------------------------------------------------------------------------

    def findValue(self,lstColumn,find,ALL=False):
        ''' Accepts list of columns and code
If ALL=False returns Found as boolean.
If ALL=True returns Found as boolean and count of instances'''

        self.lstCols=lstColumn
        self.value=find
        self.Found=False
        self.count=0
        if not ALL:
            for col in self.lstCols:
                for row in range(1,len(self.sheet[col])):
                    if self.value==self.sheet[col][row].value:
                        self.Found=True
                        return self.Found,row
                    else:
                        self.Found=False
            return [self.Found]
        else:
            for col in self.lstCols:
                for row in range(len(self.sheet[col])):
                    if self.value!=self.sheet[col][row].value:
                        self.Found=False
                        pass
                    else:
                        self.Found=True
                        self.count+=1
                        continue
                if self.count>0:
                    self.Found=True
            return [self.Found,self.count]
##----------------------------------------------------------------------------------------------------------------------------------
    def saveFile(self,folderPath=None,fileName=None):
        ''' saveFile=takes path and saves file, if no file location overwrites the current file. '''

        if folderPath==None and fileName==None:
            self.book.save(os.path.join(self.fPath,self.file))
        else:
            self.fPath=folderPath
            self.file=fileName
            self.book.save(os.path.join(self.fPath,self.file))
##----------------------------------------------------------------------------------------------------------------------------------
    def createNew(self,folder=None,file=None):
        ''' creates new excel file and return workbook obj to work on '''
        
        if folder==None and file==None:
            self.foldPath=self.fPath
            self.fileName='new.xlsx'
        elif file==None:
            self.fileName='Prem.xlsx'
            self.foldPath=folder
        else:
            self.foldPath=self.fPath
            self.fileName=file
        self.book=Workbook()
        self.ws=self.book.active
        self.book.save(os.path.join(self.foldPath,self.fileName))
        return self.book
##----------------------------------------------------------------------------------------------------------------------------------
