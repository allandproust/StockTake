from openpyxl import load_workbook # importing workbook module
fPath=r"C:\Users\arvind\Desktop\StockTake.xlsx"  # Setting File path

wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet

found=False # intialize stock found status to false

for i in range(0,10):
    scanned=input("Please scan bar scanned : ") # scanning scanned 10 times

    Add_col=ws["I"] # Column to be searched
    
    for x in range(len(Add_col)): # getting column range
        print("Finding scanned : ",x)
        if scanned==Add_col[x].value: # comparing result
            print("after comparing : ",x)
            found=True # found instance
            print("Stock found")
            if ws.cell(row=x+1, column=17).value is None: # checking total to zero
                print("Found New : ",x)
                ws.cell(row=x+1, column=16).value = scanned
                ws.cell(row=x+1, column=17).value = 1 # if zero then total = 1
            else:
                print("Found Another : ",x)
                ws.cell(row=x+1, column=16).value = scanned
                ws.cell(row=x+1, column=17).value = ws.cell(row=x+1, column=17).value+1 # if not zero total + 1
            break

print("Scanning done") #scanned ten times
print("Saving file")
wb.save(fPath)
print("You can exit now")
