from openpyxl import load_workbook # importing workbook module
fPath=r"C:\Users\MNM\Desktop\Stocktake\StockTake.xlsx"  # Setting File path

wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet

found=False # intialize stock found status to false
Add_col=ws["I"] # Column to be searched

Scan=True
def FindCode(code):  
    for x in range(len(Add_col)): # getting column range
        if code!=Add_col[x].value: # comparing result
            found=False
        else:
            print("Stock Found at : ",str(x))
            found=True
            ws.cell(row=x+1, column=16).value=None
            ws.cell(row=x+1, column=17).value=None
            ws.cell(row=x+1, column=18).value=None
            ws.cell(row=x+1, column=19).value=None
            #wb.save(fPath)
            break
    if found==False:
        print("stock not found")
for i in range(1,1000):
    try:
        if Scan==True:     
            BarCode=input("Please scan bar scanned : ") # scanning scanned 10 times
            if BarCode=="s":
                wb.save(fPath)
                break
            else:
                FindCode(BarCode)
    except KeyboardInterrupt:
        wb.Save(fPath)
        Scan==False

##for row in ws[a1:G37']
##    for cell in row:
##        cell.Value = None
