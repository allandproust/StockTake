from openpyxl import load_workbook # importing workbook module
import os

fPath=os.path.join(os.getcwd(),"StockTake.xlsx")  # Setting File path

wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet

found=False # intialize stock found status to false

for i in range(0,5):
    scanned=input("Please scan bar scanned : ") # scanning scanned 10 times

    Add_col=ws["I"] # Column to be searched
    
    for x in range(len(Add_col)): # getting column range
        if scanned!=Add_col[x].value: # comparing result
            found=False
        else:
            print("Stock Found")
            found=True
            break
    if found==False:
        print("stock not found")
    
