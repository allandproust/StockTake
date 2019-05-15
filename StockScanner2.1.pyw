from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook # importing workbook module
import winsound
import os

fPath=r"C:\Users\arvind\Desktop\StockTake.xlsx"  # Setting File path

wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet
fExcess=open(r"C:\Users\arvind\Desktop\StockExcess.txt","a")

# drawing window
window=Tk()
window.geometry("500x400")
#-----------------------------------------------
windowWidth = window.winfo_reqwidth()
windowHeight = window.winfo_reqheight()

# Gets both half the screen width/height and window width/height
positionRight = int(window.winfo_screenwidth()/2 - windowWidth)
positionDown = int(window.winfo_screenheight()/2 - windowHeight)
 
# Positions the window in the center of the page.
window.geometry("+{}+{}".format(positionRight, positionDown))
#-------------------------------------------------------------
window.title("Stock Take")
window.wm_iconbitmap(r"C:\Users\arvind\Desktop\StockFinder\FindIcon.ico")

# set the 3rd row heading value
ws.cell(row=3,column=16).value=" Bar Code"
ws.cell(row=3,column=17).value=" Found"
ws.cell(row=3,column=18).value=" Excess"
ws.cell(row=3,column=19).value=" Short"

found=False # intialize stock found status to false
saved=False # Intilized saved to false
Add_code_col=ws["I"] # Additional item code Column to be searched
item_code_col=ws["H"] # Item code Column to be searched

# setting beep sound freq
freq=440
duration=1000

# CHECK FOR STOCK MINUS
#---------------------------------------------------------------------------------------------------------------
def CheckDiff(Rw): # Checking difference of found and Total qty

    if ws.cell(row=Rw, column=19).value is None:
        ws.cell(row=Rw, column=19).value=0
        ws.cell(row=Rw, column=19).value = ws.cell(row=Rw, column=17).value - ws.cell(row=Rw, column=11).value
    else:
        ws.cell(row=Rw, column=19).value = ws.cell(row=Rw, column=17).value - ws.cell(row=Rw, column=11).value
    
#---------------------------------------------------------------------------------------------------------------

#EXCESS STOCK FOUND
#---------------------------------------------------------------------------------------------------------------
def StockStatus(code,status):
    if status=="excess":
        winsound.Beep(freq,duration) # Play warning beep
        messagebox.showinfo("Excess Stock","Stock found Excess, Please keep aside")
        lblExcess.configure(text=" Excess Scanned : "+ code)
        lblCode.configure(text="")
        fExcess.write(code+"\t1\n")
    elif status=="short":
        winsound.Beep(freq,duration) # Play warning beep
        messagebox.showinfo("Stock not Found","Stock not inwarded, Please keep aside")
        lblExcess.configure(text="Short Scanned : "+ code)
        lblCode.configure(text="")

#FUNCTION FOR FINDING CODE
#---------------------------------------------------------------------------------------------------------------
def FindCode(code):
    for Rw in range(len(Add_code_col)): # getting column range
        if code!=Add_code_col[Rw].value: #comparing result
            found=False # not found skip to last of loop
        else:
            found=True # found instance
            lblCode.configure(text="Last Scanned : "+ code) # set label text with bar code 
            if ws.cell(row=Rw+1, column=17).value is None: # checking total to zero
                ws.cell(row=Rw+1, column=16).value = code # Set bar code to column bar code
                ws.cell(row=Rw+1, column=17).value = 1 # if zero then set found column total to 1
                CheckDiff(Rw+1) # calculate minus and set column difference
            else:
                # get Qty from col=11, compare result and if excess show call StockStatus()
                if ws.cell(row=Rw+1,column=11).value != ws.cell(row=Rw+1,column=17).value:
                    ws.cell(row=Rw+1, column=16).value = code
                    ws.cell(row=Rw+1, column=17).value = ws.cell(row=Rw+1, column=17).value+1 # if not zero total + 1
                    CheckDiff(Rw+1)
                else:
                    StockStatus(code,"excess")
                    if ws.cell(row=Rw+1, column=18).value is None: # checking total to zero
                        ws.cell(row=Rw+1, column=18).value = 1 # if zero then total = 1
                    else:
                        ws.cell(row=Rw+1, column=18).value = ws.cell(row=Rw+1, column=18).value+1 # if not zero total + 1
                        CheckDiff(Rw+1)
                    lblExcessCount.configure(text="Excess Count : "+ str(ws.cell(row=Rw+1, column=18).value))
            lblScanCounter.configure(text=Rw)# Set abel
            break
    if found==False: # stock not found
        StockStatus(code,"short")
       
#---------------------------------------------------------------------------------------------------------------           
def ChkCode(): # calling main function
    txtCode.focus()
    scanned=txtCode.get()
    FindCode(scanned)
    txtCode.delete(0, END)

def SaveTake():# Saving the stocktake in between
    wb.save(fPath)
    fExcess.close()
    saved=True
    messagebox.showinfo("Save Stock Take", "File Saved ")
    
    
def QuitApp(): #Saving and completely quitting app
    SaveTake()
    messagebox.showwarning("Closing App","Stock Take successful, Thank you")
    window.quit()
    window.destroy()

window.protocol("WM_DELETE_WINDOW", QuitApp)   
# GUI layout
#---------------------------------------------------------------------------------------------------------------
lblScan=Label(window,text="Scan the Bar code: ")
lblScan.grid(column=0,row=0)

lblCode=Label(window,text="8907233000000")
lblCode.grid(column=1,row=0)

txtCode=Entry(window,width=30)
txtCode.grid(column=0,row=1)
txtCode.focus_set()

btnNext=Button(window,text="Next barcode",command=ChkCode)
btnNext.grid(column=1,row=1)

btnSave=Button(window,text="Save StockTake",command=SaveTake)
btnSave.grid(column=0,row=3)

btnQuit=Button(window,text="Quit",command=QuitApp,padx=10)
btnQuit.grid(column=1,row=3)

lblExcess=Label(window,text="Excess count : ")
lblExcess.grid(column=0,row=4)

lblExcessCount=Label(window,text="0")
lblExcessCount.grid(column=1,row=4)

lblCount=Label(window,text="Scan Count : ")
lblCount.grid(column=0,row=5)

lblScanCounter=Label(window,text="0")
lblScanCounter.grid(column=1,row=5)

#---------------------------------------------------------------------------------------------------------------

# keeping default focus on Next button
window.bind('<Return>', (lambda e, btnNext=btnNext: btnNext.invoke())) # b is your button
#window.protocol("WM_DELETE_WINDOW",QuitApp())
window.mainloop()
