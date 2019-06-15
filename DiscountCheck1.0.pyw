import openpyxl
from openpyxl import load_workbook # importing workbook module
import os,glob

yellow_fill = openpyxl.styles.colors.Color(rgb='00FFF200')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow_fill)
##cell.fill = my_fill

#----------------------------------------------------------------------------------------
print("****************")
print("CHECK DISCOUNT - Prem")
print("****************")
print("Place all barcode in \"A\" and \"B\" columns")
print("Place MRP in \"I\" column")
print("Place DISCOUNT in \"j\" column")
print("BEFORE LOADING THE FILE")
print()
print()
#-----------------------------------------------------------------------------------------

files=(glob.glob(os.path.join(os.getcwd(),"*.xlsx")))
for i in range(0,len(files)):
    print(i+1," : ",files[i])
try:
    index=int(input("Type file number to be loaded : "))
    fPath=files[index-1]
    print(fPath)
except ValueError:
    print("Exiting Program, Thank You")
    exit()
          
wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet

finished=False # for closing the program

#--------------------------------------------------------------------------------
def findCode(code): # checking code in both the columns 
    for x in range(1,ws.max_row+1):
        #row_value='A'+str(x)
        for y in "AB":
            cell_value=y+str(x)
            #print(cell_value," : ",ws[cell_value].value)
            if ws[cell_value].value==code:
                print(code," : ",ws["I"+str(x)].value," - ",ws["J"+str(x)].value)
                break
            else:
                break
#----------------------------------------------------------------------------------
while True: # scanning stock
    if finished==False:
        scan=input("Barcode : ")
        if scan=="x":
            finished==True
            wb.save(fPath)
            break
            print("You can now close the program")
        else:
            findCode(int(scan))
