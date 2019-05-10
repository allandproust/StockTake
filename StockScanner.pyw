from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook # importing workbook module

fPath=r"C:\Users\arvind\Desktop\StockTake.xlsx"  # Setting File path
wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet

# DRAWING WINDOW
window=Tk()
window.geometry('400x500')
window.title("Stock Take")

found=False # intialize stock found status to false
print("initialized Found to false")
Add_code_col=ws["I"] # Column to be searched

#FUNCTION FOR FINDING CODE
def FindCode(code):
    print("Found is false")
    print("entering loop")
    for x in range(len(Add_code_col)): # getting column range
        if code!=Add_code_col[x].value: # comparing result
            print(code," : stock not found")
            found=False
        else:
            found=True # found instance
            print(code," : stock found")
            if ws.cell(row=x+1, column=17).value is None: # checking total to zero
                ws.cell(row=x+1, column=16).value = code
                ws.cell(row=x+1, column=17).value = 1 # if zero then total = 1
                print("Writing new")
            else:
                ws.cell(row=x+1, column=16).value = code
                ws.cell(row=x+1, column=17).value = ws.cell(row=x+1, column=17).value+1 # if not zero total + 1
                print("Writing old")
            print("breaking loop")
            break
    if found==False:
        print("setting Found to false")
        messagebox.showinfo("Excess Stock","Stock found Excess, Please keep aside")
        
            
def ChkCode():
    txtCode.focus()
    scanned=txtCode.get()
    lblCode.configure(text="Scanned : "+scanned)
    FindCode(scanned)
    txtCode.delete(0, END)

def SaveTake():
    wb.save(fPath)
    messagebox.showinfo("Save Stock Take", "File Saved. You may Quit")
    
def QuitApp():
    wb.save(fPath)
    messagebox.showwarning("Closing App","Please wait while file is saved")
    window.quit()
    window.destroy()
    
    
lblScan=Label(window,text="Scan the Bar code")
lblScan.grid(column=0,row=0)

lblCode=Label(window,text="---")
lblCode.grid(column=0,row=1)

txtCode=Entry(window,width=30)
txtCode.grid(column=0,row=2)
txtCode.focus_set()

btnNext=Button(window,text="Next barcode",command=ChkCode)
btnNext.grid(column=1,row=2)

btnSave=Button(window,text="Save StockTake",command=SaveTake)
btnSave.grid(column=0,row=3)

btnQuit=Button(window,text="Quit",command=QuitApp,padx=10)
btnQuit.grid(column=1,row=3)

lblExcess=Label(window,text="---")
lblExcess.grid(column=0,row=4)

window.bind('<Return>', (lambda e, btnNext=btnNext: btnNext.invoke())) # b is your button
window.mainloop()

wb.save(fPath)

#--------------------------------------------------------------------------------------------------------------    
##    for x in range(len(Add_code_col)): # getting column range
##        print("Finding scanned : ",x)
##        if scanned==Add_code_col[x].value: # comparing result
##            print("after comparing : ",x)
##            found=True # found instance
##            print("Stock found")
##            if ws.cell(row=x+1, column=17).value is None: # checking total to zero
##                print("Found New : ",x)
##                ws.cell(row=x+1, column=16).value = scanned
##                ws.cell(row=x+1, column=17).value = 1 # if zero then total = 1
##            else:
##                print("Found Another : ",x)
##                ws.cell(row=x+1, column=16).value = scanned
##                ws.cell(row=x+1, column=17).value = ws.cell(row=x+1, column=17).value+1 # if not zero total + 1
##            break
#---------------------------------------------------------------------------------------------------------------
