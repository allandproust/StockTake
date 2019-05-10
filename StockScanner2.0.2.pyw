from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook # importing workbook module

fPath=r"C:\Users\arvind\Desktop\StockTake.xlsx"  # Setting File path
wb=load_workbook(fPath) #loading workbook in memory
ws=wb.active # setting active worksheet

# DRAWING WINDOW
window=Tk()
window.geometry('400x500')
window.title("Stock Take")

found=False # intialize stock found status to false
Add_code_col=ws["I"] # Column to be searched
Qty_col=ws["K"] # Column to be searched

#EXCESS STOCK FOUND
#---------------------------------------------------------------------------------------------------------------
def StockExcess(code,status):
    if status=="excess":
        messagebox.showinfo("Excess Stock","Stock found Excess, Please keep aside")
        lblExcess.configure(text="Scanned : "+ code)
        lblCode.configure(text="")
    elif status=="inexcess":
        messagebox.showinfo("Stock not Found","Stock not inwarded, Please keep aside")
        lblExcess.configure(text="Scanned : "+ code)
        lblCode.configure(text="")

#FUNCTION FOR FINDING CODE
#---------------------------------------------------------------------------------------------------------------
def FindCode(code):
    for x in range(len(Add_code_col)): # getting column range
        if code!=Add_code_col[x].value: # comparing result
            found=False
        else:
            found=True # found instance
            lblCode.configure(text="Scanned : "+ code)
            if ws.cell(row=x+1, column=17).value is None: # checking total to zero
                ws.cell(row=x+1, column=16).value = code
                ws.cell(row=x+1, column=17).value = 1 # if zero then total = 1
            else:
                # get Qty from col=11, compare result and if excess show call StockExcess()
                if ws.cell(row=x+1,column=11).value != ws.cell(row=x+1,column=17).value:
                    ws.cell(row=x+1, column=16).value = code
                    ws.cell(row=x+1, column=17).value = ws.cell(row=x+1, column=17).value+1 # if not zero total + 1
                else:
                    StockExcess(code,"excess")
                    if ws.cell(row=x+1, column=18).value is None: # checking total to zero
                        ws.cell(row=x+1, column=18).value = 1 # if zero then total = 1
                    else:
                        ws.cell(row=x+1, column=18).value = ws.cell(row=x+1, column=18).value+1 # if not zero total + 1
                    lblExcessCount.configure(text="Excess Count : "+ str(ws.cell(row=x+1, column=18).value))
            break
    if found==False:
        StockExcess(code,"inexcess")
#---------------------------------------------------------------------------------------------------------------           
def ChkCode(): # calling main function
    txtCode.focus()
    scanned=txtCode.get()
    FindCode(scanned)
    txtCode.delete(0, END)

def SaveTake():# Saving the stocktake in between
    wb.save(fPath)
    messagebox.showinfo("Save Stock Take", "File Saved. You may Continue or Quit")
    
def QuitApp(): #Saving and completely quitting app
    wb.save(fPath)
    messagebox.showwarning("Closing App","Please wait while file is saved")
    window.quit()
    window.destroy()
    
# GUI layout
#---------------------------------------------------------------------------------------------------------------
lblScan=Label(window,text="Scan the Bar code")
lblScan.grid(column=0,row=0)

lblCode=Label(window,text="---")
lblCode.grid(column=0,row=1)

txtCode=Entry(window,width=30)
txtCode.grid(column=0,row=2)
txtCode.focus_set()

btnNext=Button(window,text="Next barcode",command=ChkCode)
btnNext.grid(column=1,row=2)

btnSave=Button(window,text="Save StockTake",command=SaveTake)
btnSave.grid(column=0,row=3)

btnQuit=Button(window,text="Quit",command=QuitApp,padx=10)
btnQuit.grid(column=1,row=3)

lblExcess=Label(window,text="---")
lblExcess.grid(column=0,row=4)

lblExcessCount=Label(window,text="---")
lblExcessCount.grid(column=1,row=4)
#---------------------------------------------------------------------------------------------------------------

# keeping default focus on Next button
window.bind('<Return>', (lambda e, btnNext=btnNext: btnNext.invoke())) # b is your button
window.mainloop()

wb.save(fPath)
