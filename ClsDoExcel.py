from openpyxl import load_workbook
import os

class DoExcel:

    def __init__(self,folderPath,fileName): # init=opens excel file and returns file obj to work on.

        self.fPath=folderPath
        self.file=fileName
        self.book=load_workbook(os.path.join(self.fPath,self.file))
        self.sheet=self.book.active
        self.mappedCodes={}
        print("Workbook loaded")

    def __makeCol(self,num,row):  # takes in number and returns corresponding columnname
        alpha={1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M',\
               14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}
        self.colName=str(alpha.get(num,0))+str(row)
        return self.colName

    def writeHeaders(self,row,col,header):
        self.sheet.cell(row=row,column=col).value=header
        self.saveFile()
        
    def mapCodes(self): # mapCodes=takes column names in list and maps them to range.
        for cols in range(1,self.sheet.max_column+1):
            print("{} : {}".format(cols,self.sheet.cell(row=1,column=cols).value))
            self.mappedCodes[cols]=self.sheet.cell(row=1,column=cols).value
        return [self.mappedCodes,self.sheet.max_column]
            # ask for list of num
            # map them using makecol()
            # return them as mapped cols in list or tuple
    
    def setLastCols(self,count,row,*args):  # getLastcolumn=takes sheetname and number of columns,returns columnsNames in list.
        self.count=count
        headerCount=len(args)
        print("Header Count : {}".format(headerCount))
        print("Arguments Count : {}".format(len(args)))
        if headerCount==count:
            while count>0:
                for header in args:
                    self.writeHeaders(row,self.sheet.max_column+count,header)
                    count=count-1

    
    def findCode(self): # findCode=takes code and finds in the mapped range,returns row number.
        pass
    
    def writeData(self): # writeData=takes column name, row number and string to be written to cell.
        pass
    
    def saveFile(self): # saveFile=takes path and saves file, if no file location overwrites the current file.
        self.book.save(os.path.join(self.fPath,self.file))

